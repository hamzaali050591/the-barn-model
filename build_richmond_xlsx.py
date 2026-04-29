#!/usr/bin/env python3
"""
Generate Richmond Financial Model.xlsx — single-location view of the web app's
runModel engine. Source of truth for defaults: src/utils/types.ts → DEFAULT_INPUTS.
Source of truth for formulas: src/utils/engine.ts → runModel.

Differences vs the multi-location workbook:
  • numLocations locked at 1 (Richmond only)
  • openSchedule reduced to a single open month (default 4)
  • Monthly Cash Flow + Debt Schedule have one location column instead of seven
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as col
from openpyxl.workbook.defined_name import DefinedName
import os
from datetime import date

OUT = "/Users/hamzaali/Documents/The Barn/Richmond Financial Model.xlsx"

# ───── Defaults (mirror DEFAULT_INPUTS — Richmond mode) ─────
SQFT, TI_PSF, CAPEX_PSF = 9180, 35, 150
BASE_RENT_PSF, NNN_PSF = 26, 9
GP_INVESTMENT = 200_000
DEBT_PER_LOC, DEBT_RATE_PCT, DEBT_TERM_YEARS = 0, 0, 10
NUM_LOCATIONS, EXIT_MULTIPLE = 1, 3
RAMP_MONTHS, L1_HOLIDAY = 3, 3
OPEN_SCHEDULE = [4]
HOLD_MONTHS = 48
RENT_ESC, OPEX_ESC = 3, 3
BASE_RENT_ESC, NNN_ESC = 0, 2
SALARY_BASE, SALARY_STEP, PROFIT_SHARE = 84_000, 20_000, 10
REV_MODEL = "base"
PCT_OF_SALES, MIXED_BASE, MIXED_PCT = 20, 3500, 6
NON_RENT_REVENUE = 0
SPACE_LEASED_PCT = 100
RENT_INCLUDES_UTILS = True

VENDORS = [("Food Vendors", 8, 7000, 35000, True),
           ("Non-Food Vendors", 4, 5000, 25000, False)]

GAS_EQUIP = [("40lb Fryer", 120000, 0.33),
             ("2ft Flat Top Griddle", 48000, 0.30),
             ("2ft Charbroiler", 56000, 0.20),
             ("6-Burner Range", 150000, 0.23)]
GAS_HOURS, GAS_DAYS = 12, 30
GAS_COMMON = [("Commercial water heater (year-round)", 75),
              ("Space heating (annualized — peak Dec/Jan)", 175)]
GAS_RATES = (0.9, 1.05, 1.2)
GAS_SCENARIO = "mid"

ELEC_BASE = [("Double-door commercial fridge", 1.8, 0.45, 24),
             ("Double-door commercial freezer", 2.2, 0.53, 24),
             ("2 refrigerated prep tables", 1.0, 0.38, 24),
             ("Stall lighting + POS/misc", 0.5, 0.75, 12)]
ELEC_FOOD = [("Steam table / warmers", 1.5, 0.60, 12),
             ("Heat lamp", 0.5, 0.53, 12),
             ("Rice cooker", 1.0, 0.23, 12)]
ELEC_COMMON = [("HVAC (40 tons, 70% duty)", 48, 0.7, 16),
               ("Exhaust / make-up air", 1.6, 1.0, 12),
               ("Ambient lighting", 4, 1.0, 12),
               ("Hot water heater", 2, 1.0, 16),
               ("Sound system", 0.5, 1.0, 12),
               ("Restrooms (exhaust, lights, etc.)", 0.5, 1.0, 12),
               ("Security cameras / WiFi / office", 1, 1.0, 24),
               ("Overnight security lighting", 0.3, 1.0, 12)]
ELEC_RATES = (0.085, 0.10, 0.12)
ELEC_SCENARIO = "mid"

WATER_FOOD_GPD, WATER_NONFOOD_GPD, WATER_DAYS = 263, 118, 30
WATER_COMMON = [("Restrooms (~300 customers/day)", 600),
                ("Floor cleaning / mop sinks", 100),
                ("Hot water system losses / misc", 50),
                ("Outdoor hose / landscaping", 30)]
WATER_RATES = (8, 10, 12)
WATER_SCENARIO = "mid"

NON_UTIL = [
    ("Marketing", [("Social media content creation", 665),
                   ("Paid social (IG, FB, TikTok)", 1000),
                   ("Event promotion / print / signage", 415),
                   ("App marketing / push notifications", 250),
                   ("Influencer / blogger activations", 170)]),
    ("Cleaning", [("1 staff (6 hrs/day × 30 days × $15/hr)", 2700),
                  ("Cleaning supplies / restroom stock", 500)]),
    ("Grease", [("Grease trap pumping (biweekly)", 600),
                ("Grease trap hauling", 200)]),
    ("Security", [("Camera system monitoring service", 100),
                  ("On-site guard — Fri & Sat (8hr)", 0)]),
    ("Maintenance", [("HVAC preventive maintenance", 500),
                     ("Plumbing (excl. grease trap)", 200),
                     ("General repairs", 300),
                     ("Pest control", 200),
                     ("Fire suppression", 100)]),
    ("Insurance", [("General Liability", 700),
                   ("Commercial Property + Business Income", 1000),
                   ("Umbrella policy", 300)]),
    ("Technology", [("WiFi (commercial-grade)", 300),
                    ("Dashboards / analytics", 200),
                    ("Loyalty / membership platform", 200)]),
    ("Misc", [("Commercial waste / dumpster pickup", 600),
              ("Miscellaneous / contingency", 300)]),
]

# ───── Styles ─────
HDR = Font(bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="5C4033")
SUB = Font(bold=True, size=10, color="5C4033")
INPUT_FILL = PatternFill("solid", fgColor="FFF8E7")
DERIVED_FILL = PatternFill("solid", fgColor="EFEFEF")
ACCENT_FILL = PatternFill("solid", fgColor="FCE7B8")
NOTE_FONT = Font(italic=True, size=9, color="666666")

USD0 = '"$"#,##0;[Red]-"$"#,##0'
USD2 = '"$"#,##0.00;[Red]-"$"#,##0.00'
PCT1 = "0.0%"
PCT2 = "0.00%"
NUM2 = "0.00"
MULT = '0.00"x"'

def hdr(ws, row, title, span=4):
    ws.cell(row, 1, title).font = HDR
    for c in range(1, span+1):
        ws.cell(row, c).fill = HDR_FILL

def named(wb, name, ref):
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)

def setw(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[col(i)].width = w

# ───── Build workbook ─────
wb = Workbook()
wb.remove(wb.active)

# ═════ Sheet 1: Notes ═════
ws = wb.create_sheet("Notes")
setw(ws, [110])
ws['A1'] = "THE BARN — FINANCIAL MODEL"
ws['A1'].font = Font(bold=True, size=16, color="5C4033")
ws['A3'] = f"Generated: {date.today().isoformat()}"
ws['A3'].font = NOTE_FONT
notes = [
    "",
    "RICHMOND-ONLY view of the web app's runModel engine. For the 7-location",
    "portfolio model see the sister file 'The Barn — Financial Model.xlsx'.",
    "",
    "  • Source of truth for defaults: src/utils/types.ts → DEFAULT_INPUTS",
    "  • Source of truth for formulas: src/utils/engine.ts → runModel",
    "  • Mode: numLocations = 1, openSchedule = [4]",
    "",
    "All inputs live on the Inputs sheet (light-cream cells). Edit them; every other",
    "sheet recomputes via formulas. The model is fully integrated — no static values",
    "outside the Inputs sheet.",
    "",
    "Sheet roadmap:",
    "  Inputs           — every editable input + named ranges",
    "  Capital Stack    — derives TI / Debt / GP / LP / CapEx (per location)",
    "  Vendor Revenue   — vendor table + base/pct/mixed revenue split (escalating vs flat)",
    "  Gas              — equipment duty cycles + therm rates",
    "  Electric         — base + food add-on + common loads + kWh rates",
    "  Water            — vendor gallons + common loads + CCF rates",
    "  Non-Utility OpEx — line items by category (marketing, cleaning, grease, etc.)",
    "  OpEx Summary     — Year-0 monthly OpEx aggregated",
    "  Debt Schedule    — fixed-term amortization + balloon balance at exit",
    "  Monthly Cash Flow— 72-month single-location cash flow",
    "  KPIs             — IRR, MOIC, ROI, CoC, total dist, exit, equity",
    "",
    "Conventions match the web app:",
    "  • Capital call fires at max(1, openMonth − 3) — 3-month buildout window",
    "  • Annual escalators (rent / opex / base rent / NNN) apply on the open-month clock",
    "  • Senior debt amortizes on debtTermYears (fixed); balloon balance paid off from gross exit",
    "  • Exit = MAX(0, exitMultiple × T12 preCompEBITDA − debtPayoff) × (1 − profitShare%)",
    "  • Operator profit share applies to both ongoing distributions and exit",
    "  • Master lease = base rent + NNN, each escalates independently",
]
for i, line in enumerate(notes, start=4):
    ws.cell(i, 1, line)

# ═════ Sheet 2: Inputs ═════
ws = wb.create_sheet("Inputs")
setw(ws, [38, 16, 6, 38, 16])

ws['A1'] = "INPUTS — edit cream cells; named ranges drive all formulas"
ws['A1'].font = Font(bold=True, size=12, color="5C4033")

inp = [
    # (label, value, name, fmt) — col A/B
    ("CAPITAL STACK", None, None, None),  # section header
    ("Square Feet", SQFT, "sqft", "#,##0"),
    ("TI Allowance ($/PSF)", TI_PSF, "tiPSF", USD0),
    ("Base Rent ($/PSF/yr)", BASE_RENT_PSF, "baseRentPSF", USD0),
    ("NNN ($/PSF/yr)", NNN_PSF, "nnnPSF", USD0),
    ("CapEx Buildout ($/PSF)", CAPEX_PSF, "capexPSF", USD0),
    ("GP Investment ($)", GP_INVESTMENT, "gpInvestment", USD0),
    ("Debt per Location ($)", DEBT_PER_LOC, "debtPerLocation", USD0),
    ("Debt Rate (% annual)", DEBT_RATE_PCT/100, "debtRatePct", PCT1),
    ("Loan Term (years)", DEBT_TERM_YEARS, "debtTermYears", "0"),
    ("", None, None, None),
    ("REVENUE MODEL", None, None, None),
    ("Revenue Mode (base / pct / mixed)", REV_MODEL, "revenueModel", None),
    ("% of Sales Rate (%)", PCT_OF_SALES/100, "pctOfSalesRate", PCT1),
    ("Mixed Base Rent ($/vendor/mo)", MIXED_BASE, "mixedBaseRent", USD0),
    ("Mixed % of Sales Rate (%)", MIXED_PCT/100, "mixedPctRate", PCT1),
    ("Non-Rent Revenue ($/mo per loc)", NON_RENT_REVENUE, "nonRentRevenue", USD0),
    ("% Space Leased (50–100)", SPACE_LEASED_PCT/100, "spaceLeasedPct", PCT1),
    ("Rent Includes Utilities (1=yes 0=no)", 1 if RENT_INCLUDES_UTILS else 0, "rentIncludesUtilities", "0"),
    ("", None, None, None),
    ("ESCALATORS (annual, compounding, per-location clock)", None, None, None),
    ("Rent Escalator (%)", RENT_ESC/100, "rentEscalatorPct", PCT2),
    ("OpEx Escalator (%)", OPEX_ESC/100, "opexEscalatorPct", PCT2),
    ("Base Rent Escalator (%)", BASE_RENT_ESC/100, "baseRentEscalatorPct", PCT2),
    ("NNN Escalator (%)", NNN_ESC/100, "nnnEscalatorPct", PCT2),
    ("", None, None, None),
    ("OPERATOR / PROMOTE", None, None, None),
    ("Salary Base ($/yr)", SALARY_BASE, "salaryBase", USD0),
    ("Salary Step per Add'l Loc ($/yr)", SALARY_STEP, "salaryStep", USD0),
    ("Profit Share (%)", PROFIT_SHARE/100, "profitSharePct", PCT1),
    ("", None, None, None),
    ("DEAL TERMS", None, None, None),
    ("Number of Locations", NUM_LOCATIONS, "numLocations", "0"),
    ("Exit EBITDA Multiple", EXIT_MULTIPLE, "exitMultiple", MULT),
    ("Ramp Months (distribution reserve)", RAMP_MONTHS, "rampMonths", "0"),
    ("L1 Lease Holiday (months)", L1_HOLIDAY, "l1LeaseHolidayMonths", "0"),
    ("Hold Period (months)", HOLD_MONTHS, "holdMonths", "0"),
]
row = 3
for label, value, name, fmt in inp:
    ws.cell(row, 1, label)
    if value is None and name is None:
        ws.cell(row, 1).font = SUB if label else Font()
    else:
        ws.cell(row, 2, value)
        ws.cell(row, 2).fill = INPUT_FILL
        if fmt:
            ws.cell(row, 2).number_format = fmt
        if name:
            named(wb, name, f"Inputs!$B${row}")
    row += 1

# Open schedule (as a table on the right)
ws.cell(3, 4, "OPEN SCHEDULE (month each location opens)").font = SUB
for i, om in enumerate(OPEN_SCHEDULE):
    r = 4 + i
    ws.cell(r, 4, f"L{i+1} Open Month")
    ws.cell(r, 5, om)
    ws.cell(r, 5).fill = INPUT_FILL
    named(wb, f"openMonth_{i+1}", f"Inputs!$E${r}")

# Vendor table on Inputs
ws.cell(13, 4, "VENDOR CATEGORIES").font = SUB
ws.cell(14, 4, "Name"); ws.cell(14, 5, "Count"); ws.cell(14, 4).font = SUB; ws.cell(14, 5).font = SUB
ws.cell(14, 4).fill = HDR_FILL; ws.cell(14, 5).fill = HDR_FILL
ws.cell(14, 4).font = HDR; ws.cell(14, 5).font = HDR
for i, (name, count, rent, sales, isFood) in enumerate(VENDORS):
    r = 15 + i
    ws.cell(r, 4, name)
    ws.cell(r, 5, count); ws.cell(r, 5).fill = INPUT_FILL
ws.cell(18, 4, "Food Rent ($/mo)"); ws.cell(18, 5, VENDORS[0][2]); ws.cell(18, 5).fill = INPUT_FILL; ws.cell(18, 5).number_format = USD0
named(wb, "foodRent", "Inputs!$E$18")
ws.cell(19, 4, "Food Sales ($/mo)"); ws.cell(19, 5, VENDORS[0][3]); ws.cell(19, 5).fill = INPUT_FILL; ws.cell(19, 5).number_format = USD0
named(wb, "foodSales", "Inputs!$E$19")
ws.cell(20, 4, "Non-Food Rent ($/mo)"); ws.cell(20, 5, VENDORS[1][2]); ws.cell(20, 5).fill = INPUT_FILL; ws.cell(20, 5).number_format = USD0
named(wb, "nonFoodRent", "Inputs!$E$20")
ws.cell(21, 4, "Non-Food Sales ($/mo)"); ws.cell(21, 5, VENDORS[1][3]); ws.cell(21, 5).fill = INPUT_FILL; ws.cell(21, 5).number_format = USD0
named(wb, "nonFoodSales", "Inputs!$E$21")

# Named ranges for vendor counts
named(wb, "numFoodVendors", "Inputs!$E$15")
named(wb, "numNonFoodVendors", "Inputs!$E$16")

# ═════ Sheet 3: Capital Stack ═════
ws = wb.create_sheet("Capital Stack")
setw(ws, [38, 18])
ws['A1'] = "CAPITAL STACK (per location)"; ws['A1'].font = Font(bold=True, size=12, color="5C4033")
rows = [
    ("Total CapEx", "=sqft*capexPSF", USD0),
    ("Total TI (DPEG)", "=sqft*tiPSF", USD0),
    ("Equity Slot (CapEx − TI)", "=MAX(0,B2-B3)", USD0),
    ("GP Used (clamped)", "=MIN(gpInvestment,B4)", USD0),
    ("Debt Slot (Equity − GP)", "=MAX(0,B4-B5)", USD0),
    ("Debt Used (clamped)", "=MIN(MAX(0,debtPerLocation),B6)", USD0),
    ("Investor Equity (LP+GP) per loc", "=MAX(0,B4-B7)", USD0),
    ("LP Investment per loc", "=MAX(0,B8-B5)", USD0),
    ("Base Rent ($/mo per loc)", "=baseRentPSF*sqft/12", USD0),
    ("NNN ($/mo per loc)", "=nnnPSF*sqft/12", USD0),
    ("Total Master Lease ($/mo per loc, Year 0)", "=B10+B11", USD0),
]
for i, (label, formula, fmt) in enumerate(rows):
    r = i + 2
    ws.cell(r, 1, label)
    ws.cell(r, 2, formula)
    ws.cell(r, 2).fill = DERIVED_FILL
    ws.cell(r, 2).number_format = fmt

# Named ranges to feed downstream sheets
named(wb, "totalCapex", "'Capital Stack'!$B$2")
named(wb, "tiTotal", "'Capital Stack'!$B$3")
named(wb, "gpUsed", "'Capital Stack'!$B$5")
named(wb, "debtUsed", "'Capital Stack'!$B$7")
named(wb, "investorEquityPerLoc", "'Capital Stack'!$B$8")
named(wb, "lpPerLoc", "'Capital Stack'!$B$9")
named(wb, "monthlyBaseRentPerLoc", "'Capital Stack'!$B$10")
named(wb, "monthlyNnnPerLoc", "'Capital Stack'!$B$11")

# ═════ Sheet 4: Vendor Revenue ═════
ws = wb.create_sheet("Vendor Revenue")
setw(ws, [38, 14, 14, 14])
ws['A1'] = "VENDOR REVENUE — base / pct / mixed split"; ws['A1'].font = Font(bold=True, size=12, color="5C4033")
ws.cell(3, 1, "Total Vendors")
ws.cell(3, 2, "=numFoodVendors+numNonFoodVendors")
ws.cell(4, 1, "Base Rent Total (Food×rent + NonFood×rent)")
ws.cell(4, 2, "=numFoodVendors*foodRent+numNonFoodVendors*nonFoodRent"); ws.cell(4, 2).number_format = USD0
ws.cell(5, 1, "Sales Total (Food×sales + NonFood×sales)")
ws.cell(5, 2, "=numFoodVendors*foodSales+numNonFoodVendors*nonFoodSales"); ws.cell(5, 2).number_format = USD0
ws.cell(7, 1, "Escalating Rent — pre-leased-factor").font = SUB
ws.cell(7, 2, '=IF(revenueModel="base",B4,IF(revenueModel="pct",0,B3*mixedBaseRent))')
ws.cell(7, 2).number_format = USD0
ws.cell(8, 1, "Flat Rent — pre-leased-factor").font = SUB
ws.cell(8, 2, '=IF(revenueModel="base",0,IF(revenueModel="pct",B5*pctOfSalesRate,B5*mixedPctRate))')
ws.cell(8, 2).number_format = USD0

ws.cell(10, 1, "Escalating Rent (× spaceLeasedPct)").font = SUB
ws.cell(10, 2, "=B7*spaceLeasedPct"); ws.cell(10, 2).number_format = USD0; ws.cell(10, 2).fill = ACCENT_FILL
ws.cell(11, 1, "Flat Rent (× spaceLeasedPct)").font = SUB
ws.cell(11, 2, "=B8*spaceLeasedPct"); ws.cell(11, 2).number_format = USD0; ws.cell(11, 2).fill = ACCENT_FILL
ws.cell(12, 1, "Total Monthly Vendor Rent (Year 0)")
ws.cell(12, 2, "=B10+B11"); ws.cell(12, 2).number_format = USD0; ws.cell(12, 2).fill = DERIVED_FILL
ws.cell(13, 1, "Non-Rent Revenue (per loc, Year 0)")
ws.cell(13, 2, "=nonRentRevenue"); ws.cell(13, 2).number_format = USD0; ws.cell(13, 2).fill = ACCENT_FILL
ws.cell(14, 1, "Total Revenue (per loc, Year 0)")
ws.cell(14, 2, "=B12+B13"); ws.cell(14, 2).number_format = USD0; ws.cell(14, 2).fill = DERIVED_FILL

named(wb, "escalatingRent", "'Vendor Revenue'!$B$10")
named(wb, "flatRent", "'Vendor Revenue'!$B$11")

# ═════ Sheet 5: Gas ═════
ws = wb.create_sheet("Gas")
setw(ws, [40, 14, 14, 14])
ws['A1'] = "GAS — per food vendor"; ws['A1'].font = Font(bold=True, size=12, color="5C4033")
ws.cell(3, 1, "Equipment").font = HDR; ws.cell(3, 2, "BTU/hr (Max)").font = HDR
ws.cell(3, 3, "Duty Cycle").font = HDR; ws.cell(3, 4, "Effective BTU/hr").font = HDR
for i, (n, b, d) in enumerate(GAS_EQUIP):
    r = 4 + i
    ws.cell(r, 1, n); ws.cell(r, 2, b); ws.cell(r, 2).fill = INPUT_FILL
    ws.cell(r, 3, d); ws.cell(r, 3).fill = INPUT_FILL
    ws.cell(r, 4, f"=B{r}*C{r}"); ws.cell(r, 4).number_format = "#,##0"
sum_row = 4 + len(GAS_EQUIP)
ws.cell(sum_row, 1, "Total Effective BTU/hr").font = SUB
ws.cell(sum_row, 4, f"=SUM(D4:D{sum_row-1})"); ws.cell(sum_row, 4).number_format = "#,##0"

ws.cell(sum_row+2, 1, "Hours per Day"); ws.cell(sum_row+2, 2, GAS_HOURS); ws.cell(sum_row+2, 2).fill = INPUT_FILL
ws.cell(sum_row+3, 1, "Days per Month"); ws.cell(sum_row+3, 2, GAS_DAYS); ws.cell(sum_row+3, 2).fill = INPUT_FILL
ws.cell(sum_row+4, 1, "Monthly Therms (per food vendor)")
ws.cell(sum_row+4, 2, f"=D{sum_row}*B{sum_row+2}*B{sum_row+3}/100000"); ws.cell(sum_row+4, 2).number_format = NUM2
named(wb, "gasMonthlyTherms", f"Gas!$B${sum_row+4}")

ws.cell(sum_row+6, 1, "Scenario Rates ($/therm)").font = SUB
ws.cell(sum_row+7, 1, "Low"); ws.cell(sum_row+7, 2, GAS_RATES[0]); ws.cell(sum_row+7, 2).fill = INPUT_FILL; ws.cell(sum_row+7, 2).number_format = USD2
ws.cell(sum_row+8, 1, "Mid"); ws.cell(sum_row+8, 2, GAS_RATES[1]); ws.cell(sum_row+8, 2).fill = INPUT_FILL; ws.cell(sum_row+8, 2).number_format = USD2
ws.cell(sum_row+9, 1, "High"); ws.cell(sum_row+9, 2, GAS_RATES[2]); ws.cell(sum_row+9, 2).fill = INPUT_FILL; ws.cell(sum_row+9, 2).number_format = USD2
ws.cell(sum_row+10, 1, "Active Scenario (low/mid/high)"); ws.cell(sum_row+10, 2, GAS_SCENARIO); ws.cell(sum_row+10, 2).fill = INPUT_FILL
named(wb, "gasScenario", f"Gas!$B${sum_row+10}")
ws.cell(sum_row+11, 1, "Active Rate ($/therm)")
ws.cell(sum_row+11, 2, f'=IF(gasScenario="low",B{sum_row+7},IF(gasScenario="high",B{sum_row+9},B{sum_row+8}))')
ws.cell(sum_row+11, 2).number_format = USD2
named(wb, "gasRate", f"Gas!$B${sum_row+11}")

ws.cell(sum_row+13, 1, "Per Food Vendor / mo (gas cost)").font = SUB
ws.cell(sum_row+13, 2, f"=B{sum_row+4}*B{sum_row+11}"); ws.cell(sum_row+13, 2).number_format = USD0; ws.cell(sum_row+13, 2).fill = DERIVED_FILL
named(wb, "gasFoodVendor", f"Gas!$B${sum_row+13}")

# Common-area gas (water heater + space heating)
ws.cell(sum_row+15, 1, "COMMON AREA GAS — building-level (Barn pays)").font = SUB
ws.cell(sum_row+16, 1, "Load").font = HDR; ws.cell(sum_row+16, 2, "Therms / mo").font = HDR; ws.cell(sum_row+16, 3, "Monthly $").font = HDR
for i, (n, therms) in enumerate(GAS_COMMON):
    r = sum_row + 17 + i
    ws.cell(r, 1, n)
    ws.cell(r, 2, therms); ws.cell(r, 2).fill = INPUT_FILL
    ws.cell(r, 3, f"=B{r}*gasRate"); ws.cell(r, 3).number_format = USD2
common_end_row = sum_row + 17 + len(GAS_COMMON)
ws.cell(common_end_row, 1, "Total Common Area Therms / mo").font = SUB
ws.cell(common_end_row, 2, f"=SUM(B{sum_row+17}:B{common_end_row-1})"); ws.cell(common_end_row, 2).fill = DERIVED_FILL
ws.cell(common_end_row, 3, f"=SUM(C{sum_row+17}:C{common_end_row-1})"); ws.cell(common_end_row, 3).number_format = USD0; ws.cell(common_end_row, 3).fill = ACCENT_FILL
named(wb, "gasCommon", f"Gas!$C${common_end_row}")

# ═════ Sheet 6: Electric ═════
ws = wb.create_sheet("Electric")
setw(ws, [40, 12, 12, 12, 12, 14])
ws['A1'] = "ELECTRIC — per vendor & common"; ws['A1'].font = Font(bold=True, size=12, color="5C4033")
def write_load_block(ws, start_row, title, loads):
    ws.cell(start_row, 1, title).font = SUB
    ws.cell(start_row+1, 1, "Equipment").font = HDR
    headers = ["kW", "Duty", "Eff kW", "Hrs/Day", "kWh/mo"]
    for j, h in enumerate(headers):
        ws.cell(start_row+1, 2+j, h).font = HDR
    for i, (n, kw, duty, hrs) in enumerate(loads):
        r = start_row + 2 + i
        ws.cell(r, 1, n)
        ws.cell(r, 2, kw); ws.cell(r, 2).fill = INPUT_FILL
        ws.cell(r, 3, duty); ws.cell(r, 3).fill = INPUT_FILL
        ws.cell(r, 4, f"=B{r}*C{r}"); ws.cell(r, 4).number_format = NUM2
        ws.cell(r, 5, hrs); ws.cell(r, 5).fill = INPUT_FILL
        ws.cell(r, 6, f"=D{r}*E{r}*30"); ws.cell(r, 6).number_format = "#,##0"
    sub_r = start_row + 2 + len(loads)
    ws.cell(sub_r, 1, "Subtotal kWh/mo").font = SUB
    ws.cell(sub_r, 6, f"=SUM(F{start_row+2}:F{sub_r-1})"); ws.cell(sub_r, 6).number_format = "#,##0"
    return sub_r

base_end = write_load_block(ws, 3, "BASE LOAD (per vendor)", ELEC_BASE)
food_end = write_load_block(ws, base_end+2, "FOOD ADD-ONS (per food vendor)", ELEC_FOOD)
common_end = write_load_block(ws, food_end+2, "COMMON AREA (Barn-wide)", ELEC_COMMON)

# Roll-up
r = common_end + 2
ws.cell(r, 1, "Per Food Vendor kWh/mo"); ws.cell(r, 6, f"=F{base_end}+F{food_end}"); ws.cell(r, 6).number_format = "#,##0"
ws.cell(r+1, 1, "Per Non-Food Vendor kWh/mo"); ws.cell(r+1, 6, f"=F{base_end}"); ws.cell(r+1, 6).number_format = "#,##0"
ws.cell(r+2, 1, "Common Area kWh/mo"); ws.cell(r+2, 6, f"=F{common_end}"); ws.cell(r+2, 6).number_format = "#,##0"

# Rates
rate_r = r + 4
ws.cell(rate_r, 1, "Rates ($/kWh)").font = SUB
ws.cell(rate_r+1, 1, "Low"); ws.cell(rate_r+1, 2, ELEC_RATES[0]); ws.cell(rate_r+1, 2).fill = INPUT_FILL; ws.cell(rate_r+1, 2).number_format = '"$"0.000'
ws.cell(rate_r+2, 1, "Mid"); ws.cell(rate_r+2, 2, ELEC_RATES[1]); ws.cell(rate_r+2, 2).fill = INPUT_FILL; ws.cell(rate_r+2, 2).number_format = '"$"0.000'
ws.cell(rate_r+3, 1, "High"); ws.cell(rate_r+3, 2, ELEC_RATES[2]); ws.cell(rate_r+3, 2).fill = INPUT_FILL; ws.cell(rate_r+3, 2).number_format = '"$"0.000'
ws.cell(rate_r+4, 1, "Active Scenario"); ws.cell(rate_r+4, 2, ELEC_SCENARIO); ws.cell(rate_r+4, 2).fill = INPUT_FILL
named(wb, "elecScenario", f"Electric!$B${rate_r+4}")
ws.cell(rate_r+5, 1, "Active Rate")
ws.cell(rate_r+5, 2, f'=IF(elecScenario="low",B{rate_r+1},IF(elecScenario="high",B{rate_r+3},B{rate_r+2}))')
ws.cell(rate_r+5, 2).number_format = '"$"0.000'
named(wb, "elecRate", f"Electric!$B${rate_r+5}")

# Cost roll-ups
roll_r = rate_r + 7
ws.cell(roll_r, 1, "Per Food Vendor / mo"); ws.cell(roll_r, 2, f"=F{r}*elecRate"); ws.cell(roll_r, 2).number_format = USD0; ws.cell(roll_r, 2).fill = DERIVED_FILL
named(wb, "elecFoodVendor", f"Electric!$B${roll_r}")
ws.cell(roll_r+1, 1, "Per Non-Food Vendor / mo"); ws.cell(roll_r+1, 2, f"=F{r+1}*elecRate"); ws.cell(roll_r+1, 2).number_format = USD0; ws.cell(roll_r+1, 2).fill = DERIVED_FILL
named(wb, "elecNonFoodVendor", f"Electric!$B${roll_r+1}")
ws.cell(roll_r+2, 1, "Common Area / mo"); ws.cell(roll_r+2, 2, f"=F{r+2}*elecRate"); ws.cell(roll_r+2, 2).number_format = USD0; ws.cell(roll_r+2, 2).fill = DERIVED_FILL
named(wb, "elecCommon", f"Electric!$B${roll_r+2}")

# ═════ Sheet 7: Water ═════
ws = wb.create_sheet("Water")
setw(ws, [40, 14, 14, 14, 14])
ws['A1'] = "WATER / SEWER — vendor & common"; ws['A1'].font = Font(bold=True, size=12, color="5C4033")
ws.cell(3, 1, "Per Food Vendor — gal/day"); ws.cell(3, 2, WATER_FOOD_GPD); ws.cell(3, 2).fill = INPUT_FILL
ws.cell(4, 1, "Per Non-Food Vendor — gal/day"); ws.cell(4, 2, WATER_NONFOOD_GPD); ws.cell(4, 2).fill = INPUT_FILL
ws.cell(5, 1, "Days per Month"); ws.cell(5, 2, WATER_DAYS); ws.cell(5, 2).fill = INPUT_FILL

ws.cell(7, 1, "Common Area Loads (gal/day)").font = SUB
for i, (n, gpd) in enumerate(WATER_COMMON):
    r = 8 + i
    ws.cell(r, 1, n); ws.cell(r, 2, gpd); ws.cell(r, 2).fill = INPUT_FILL
common_end_w = 8 + len(WATER_COMMON)
ws.cell(common_end_w, 1, "Total Common gal/day").font = SUB
ws.cell(common_end_w, 2, f"=SUM(B8:B{common_end_w-1})")

# Rates
ws.cell(common_end_w+2, 1, "Rates ($/CCF)").font = SUB
ws.cell(common_end_w+3, 1, "Low"); ws.cell(common_end_w+3, 2, WATER_RATES[0]); ws.cell(common_end_w+3, 2).fill = INPUT_FILL; ws.cell(common_end_w+3, 2).number_format = USD2
ws.cell(common_end_w+4, 1, "Mid"); ws.cell(common_end_w+4, 2, WATER_RATES[1]); ws.cell(common_end_w+4, 2).fill = INPUT_FILL; ws.cell(common_end_w+4, 2).number_format = USD2
ws.cell(common_end_w+5, 1, "High"); ws.cell(common_end_w+5, 2, WATER_RATES[2]); ws.cell(common_end_w+5, 2).fill = INPUT_FILL; ws.cell(common_end_w+5, 2).number_format = USD2
ws.cell(common_end_w+6, 1, "Active Scenario"); ws.cell(common_end_w+6, 2, WATER_SCENARIO); ws.cell(common_end_w+6, 2).fill = INPUT_FILL
named(wb, "waterScenario", f"Water!$B${common_end_w+6}")
ws.cell(common_end_w+7, 1, "Active Rate ($/CCF)")
ws.cell(common_end_w+7, 2, f'=IF(waterScenario="low",B{common_end_w+3},IF(waterScenario="high",B{common_end_w+5},B{common_end_w+4}))')
ws.cell(common_end_w+7, 2).number_format = USD2
named(wb, "waterRate", f"Water!$B${common_end_w+7}")

# Cost roll-ups
roll = common_end_w + 9
ws.cell(roll, 1, "Per Food Vendor / mo"); ws.cell(roll, 2, f"=B3*B5/748*waterRate"); ws.cell(roll, 2).number_format = USD0; ws.cell(roll, 2).fill = DERIVED_FILL
named(wb, "waterFoodVendor", f"Water!$B${roll}")
ws.cell(roll+1, 1, "Per Non-Food Vendor / mo"); ws.cell(roll+1, 2, f"=B4*B5/748*waterRate"); ws.cell(roll+1, 2).number_format = USD0; ws.cell(roll+1, 2).fill = DERIVED_FILL
named(wb, "waterNonFoodVendor", f"Water!$B${roll+1}")
ws.cell(roll+2, 1, "Common Area / mo"); ws.cell(roll+2, 2, f"=B{common_end_w}*B5/748*waterRate"); ws.cell(roll+2, 2).number_format = USD0; ws.cell(roll+2, 2).fill = DERIVED_FILL
named(wb, "waterCommon", f"Water!$B${roll+2}")

# ═════ Sheet 8: Non-Utility OpEx ═════
ws = wb.create_sheet("Non-Utility OpEx")
setw(ws, [40, 14])
ws['A1'] = "NON-UTILITY OPEX — line items"; ws['A1'].font = Font(bold=True, size=12, color="5C4033")
r = 3
section_totals = []
for cat, items in NON_UTIL:
    ws.cell(r, 1, cat).font = SUB
    r += 1
    start = r
    for n, m in items:
        ws.cell(r, 1, n); ws.cell(r, 2, m); ws.cell(r, 2).fill = INPUT_FILL; ws.cell(r, 2).number_format = USD0
        r += 1
    ws.cell(r, 1, f"{cat} subtotal").font = SUB
    ws.cell(r, 2, f"=SUM(B{start}:B{r-1})"); ws.cell(r, 2).number_format = USD0; ws.cell(r, 2).fill = DERIVED_FILL
    section_totals.append(r)
    r += 2
ws.cell(r, 1, "TOTAL Non-Utility OpEx / mo").font = Font(bold=True)
sums = "+".join(f"B{tr}" for tr in section_totals)
ws.cell(r, 2, f"={sums}"); ws.cell(r, 2).number_format = USD0; ws.cell(r, 2).fill = ACCENT_FILL
named(wb, "nonUtilityTotal", f"'Non-Utility OpEx'!$B${r}")

# ═════ Sheet 9: OpEx Summary ═════
ws = wb.create_sheet("OpEx Summary")
setw(ws, [42, 14])
ws['A1'] = "MONTHLY OPEX SUMMARY (Year 0, per location)"; ws['A1'].font = Font(bold=True, size=12, color="5C4033")
rows = [
    ("Food Vendor utility / mo", "=gasFoodVendor+elecFoodVendor+waterFoodVendor"),
    ("Non-Food Vendor utility / mo", "=elecNonFoodVendor+waterNonFoodVendor"),
    ("# Food Vendors × cost", "=numFoodVendors*B2"),
    ("# Non-Food Vendors × cost", "=numNonFoodVendors*B3"),
    ("Vendor Utilities (if rentIncludesUtilities)", "=IF(rentIncludesUtilities=1,B4+B5,0)"),
    ("Common Area Utilities", "=gasCommon+elecCommon+waterCommon"),
    ("Non-Utility OpEx", "=nonUtilityTotal"),
    ("TOTAL Monthly OpEx (Year 0, per loc)", "=B6+B7+B8"),
]
for i, (label, formula) in enumerate(rows):
    r = i + 2
    ws.cell(r, 1, label)
    ws.cell(r, 2, formula); ws.cell(r, 2).number_format = USD0
    if i == len(rows)-1:
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2).fill = ACCENT_FILL
    else:
        ws.cell(r, 2).fill = DERIVED_FILL
named(wb, "monthlyOpexPerLoc", "'OpEx Summary'!$B$9")

# ═════ Sheet 10: Debt Schedule ═════
ws = wb.create_sheet("Debt Schedule")
setw(ws, [40, 16])
ws['A1'] = "DEBT SCHEDULE — fixed amortization term; balloon at exit if term > hold"; ws['A1'].font = Font(bold=True, size=12, color="5C4033")
ws.cell(3, 1, "Metric").font = HDR
for j in range(1, 2):
    ws.cell(3, j+1, f"L{j}").font = HDR; ws.cell(3, j+1).fill = HDR_FILL; ws.cell(3, j+1).font = HDR
ws.cell(4, 1, "Open Month")
ws.cell(5, 1, "Capital Call Month = MAX(1, Open−3)")
ws.cell(6, 1, "Loan Term (mo) = debtTermYears × 12")
ws.cell(7, 1, "Active in Stack (j ≤ numLocations)")
ws.cell(8, 1, "Monthly Rate")
ws.cell(9, 1, "Level Monthly P&I")
ws.cell(10, 1, "Payments Made by holdMonths")
ws.cell(11, 1, "Loan Balance @ Exit (balloon)")
for j in range(1, 2):
    c = j + 1
    ws.cell(4, c, f"=openMonth_{j}")
    ws.cell(5, c, f"=MAX(1,openMonth_{j}-3)")
    ws.cell(6, c, f"=MAX(1,debtTermYears*12)")
    ws.cell(7, c, f"=IF({j}<=numLocations,1,0)")
    ws.cell(8, c, "=debtRatePct/12")
    ws.cell(8, c).number_format = "0.00%"
    ws.cell(9, c, f'=IF(OR(debtUsed=0,{col(c)}6=0,{col(c)}7=0),0,IF({col(c)}8=0,debtUsed/{col(c)}6,debtUsed*({col(c)}8*(1+{col(c)}8)^{col(c)}6)/((1+{col(c)}8)^{col(c)}6-1)))')
    ws.cell(9, c).number_format = USD2
    ws.cell(10, c, f'=IF({col(c)}7=0,0,MIN({col(c)}6,MAX(0,holdMonths-{col(c)}5+1)))')
    ws.cell(10, c).number_format = "0"
    ws.cell(11, c,
        f'=IF(OR({col(c)}7=0,debtUsed=0,{col(c)}6=0,{col(c)}5>holdMonths),0,'
        f'IF({col(c)}8=0,'
        f'MAX(0,debtUsed*(1-{col(c)}10/{col(c)}6)),'
        f'MAX(0,debtUsed*((1+{col(c)}8)^{col(c)}6-(1+{col(c)}8)^{col(c)}10)/((1+{col(c)}8)^{col(c)}6-1))))')
    ws.cell(11, c).number_format = USD0

# Named ranges for the per-location lookups (used in Monthly sheet)
for j in range(1, 2):
    c = j + 1
    named(wb, f"openMo_{j}", f"'Debt Schedule'!${col(c)}$4")
    named(wb, f"callMo_{j}", f"'Debt Schedule'!${col(c)}$5")
    named(wb, f"loanTerm_{j}", f"'Debt Schedule'!${col(c)}$6")
    named(wb, f"locActive_{j}", f"'Debt Schedule'!${col(c)}$7")
    named(wb, f"monthlyRate_{j}", f"'Debt Schedule'!${col(c)}$8")
    named(wb, f"levelPmt_{j}", f"'Debt Schedule'!${col(c)}$9")
    named(wb, f"loanBalAtExit_{j}", f"'Debt Schedule'!${col(c)}$11")

# Total debt payoff (balloon) at exit — Richmond is 1 location, but the named
# range is a row to keep the exit formula identical across portfolio + Richmond.
named(wb, "debtPayoff", "'Debt Schedule'!$B$11:$B$11")

# ═════ Sheet 11: Monthly Cash Flow ═════
ws = wb.create_sheet("Monthly Cash Flow")
NLOC = 1

# Column plan:
# A = month
# B..H = j-Active (1-7)
# I..O = j-Distributing
# P..V = j-yearsOpen
# W..AC = j-revenue
# AD..AJ = j-opex
# AK..AQ = j-lease
# AR..AX = j-debtService
# AY = sumActive (nActive)
# AZ = sumDistributing
# BA = totalRevenue
# BB = totalOpex
# BC = totalLease
# BD = preCompEBITDA
# BE = salaryYearsOpen
# BF = corpSalary
# BG = postSalaryEBITDA
# BH = totalDebtService
# BI = postDebtEBITDA
# BJ = distributableNOI
# BK = profitShare
# BL = distributions
# BM = capitalCall
# BN = exitProceeds
# BO = exitProfitShare
# BP = netCashFlow

headers = (
    ["Month"] +
    [f"j{j}-Active" for j in range(1, NLOC+1)] +
    [f"j{j}-Dist" for j in range(1, NLOC+1)] +
    [f"j{j}-YearsOpen" for j in range(1, NLOC+1)] +
    [f"j{j}-Revenue" for j in range(1, NLOC+1)] +
    [f"j{j}-OpEx" for j in range(1, NLOC+1)] +
    [f"j{j}-Lease" for j in range(1, NLOC+1)] +
    [f"j{j}-DebtSvc" for j in range(1, NLOC+1)] +
    ["nActive", "nDistrib", "Revenue", "OpEx", "Lease",
     "preCompEBITDA", "SalYrsOpen", "CorpSalary", "postSalaryEBITDA",
     "DebtService", "postDebtEBITDA", "DistributableNOI",
     "ProfitShare", "Distributions", "CapitalCall",
     "ExitProceeds", "ExitProfitShare", "NetCashFlow"]
)

for i, h in enumerate(headers, start=1):
    ws.cell(1, i, h).font = HDR
    ws.cell(1, i).fill = HDR_FILL

# Column letter helpers
def cl(idx): return col(idx)

ACTIVE_OFF = 2  # B
DIST_OFF = ACTIVE_OFF + NLOC  # I
YR_OFF = DIST_OFF + NLOC  # P
REV_OFF = YR_OFF + NLOC  # W
OPEX_OFF = REV_OFF + NLOC  # AD
LEASE_OFF = OPEX_OFF + NLOC  # AK
DEBT_OFF = LEASE_OFF + NLOC  # AR

NACTIVE = DEBT_OFF + NLOC
NDIST = NACTIVE + 1
TREV = NDIST + 1
TOPEX = TREV + 1
TLEASE = TOPEX + 1
PRECOMP = TLEASE + 1
SALYR = PRECOMP + 1
CORPSAL = SALYR + 1
POSTSAL = CORPSAL + 1
TDEBT = POSTSAL + 1
POSTDEBT = TDEBT + 1
DISTNOI = POSTDEBT + 1
PSHARE = DISTNOI + 1
DIST = PSHARE + 1
CALL = DIST + 1
EXIT = CALL + 1
EXITP = EXIT + 1
NETCF = EXITP + 1

MAX_MONTHS = 72
for m in range(1, MAX_MONTHS + 1):
    r = m + 1
    ws.cell(r, 1, m)
    # Per-location columns
    for j in range(1, NLOC+1):
        # j-Active = locActive_j AND m >= openMonth_j AND m <= holdMonths
        ws.cell(r, ACTIVE_OFF + j - 1,
                f"=IF(AND(locActive_{j}=1,A{r}>=openMo_{j},A{r}<=holdMonths),1,0)")
        # j-Dist = j-Active AND m - rampMonths >= openMo_j
        ws.cell(r, DIST_OFF + j - 1,
                f"=IF(AND({cl(ACTIVE_OFF+j-1)}{r}=1,A{r}-rampMonths>=openMo_{j}),1,0)")
        # j-YearsOpen = IF(active, FLOOR((m - openMo)/12), 0)
        ws.cell(r, YR_OFF + j - 1,
                f"=IF({cl(ACTIVE_OFF+j-1)}{r}=1,FLOOR((A{r}-openMo_{j})/12,1),0)")
        # j-Revenue = active * (escalatingRent * (1+rentEsc)^yrs + flatRent + nonRentRevenue * (1+rentEsc)^yrs)
        ws.cell(r, REV_OFF + j - 1,
                f"=IF({cl(ACTIVE_OFF+j-1)}{r}=1,(escalatingRent+nonRentRevenue)*(1+rentEscalatorPct)^{cl(YR_OFF+j-1)}{r}+flatRent,0)")
        # j-OpEx = active * (monthlyOpex * (1+opexEsc)^yrs)
        ws.cell(r, OPEX_OFF + j - 1,
                f"=IF({cl(ACTIVE_OFF+j-1)}{r}=1,monthlyOpexPerLoc*(1+opexEscalatorPct)^{cl(YR_OFF+j-1)}{r},0)")
        # j-Lease = active * (baseRent*(1+baseRentEsc)^yrs + nnn*(1+nnnEsc)^yrs)
        # L1 holiday zeroes the full lease (base + NNN) for the first N months.
        if j == 1:
            ws.cell(r, LEASE_OFF + j - 1,
                    f"=IF({cl(ACTIVE_OFF+j-1)}{r}=0,0,IF(AND(A{r}>=openMo_1,A{r}<openMo_1+l1LeaseHolidayMonths),0,monthlyBaseRentPerLoc*(1+baseRentEscalatorPct)^{cl(YR_OFF+j-1)}{r}+monthlyNnnPerLoc*(1+nnnEscalatorPct)^{cl(YR_OFF+j-1)}{r}))")
        else:
            ws.cell(r, LEASE_OFF + j - 1,
                    f"=IF({cl(ACTIVE_OFF+j-1)}{r}=1,monthlyBaseRentPerLoc*(1+baseRentEscalatorPct)^{cl(YR_OFF+j-1)}{r}+monthlyNnnPerLoc*(1+nnnEscalatorPct)^{cl(YR_OFF+j-1)}{r},0)")
        # j-DebtSvc: payment if m >= callMo and m < callMo + term and j is in stack
        ws.cell(r, DEBT_OFF + j - 1,
                f"=IF(AND(locActive_{j}=1,A{r}>=callMo_{j},A{r}<callMo_{j}+loanTerm_{j},A{r}<=holdMonths),levelPmt_{j},0)")

    # Aggregations
    rng = lambda off: f"{cl(off)}{r}:{cl(off+NLOC-1)}{r}"
    ws.cell(r, NACTIVE, f"=SUM({rng(ACTIVE_OFF)})")
    ws.cell(r, NDIST,   f"=SUM({rng(DIST_OFF)})")
    ws.cell(r, TREV,    f"=SUM({rng(REV_OFF)})")
    ws.cell(r, TOPEX,   f"=SUM({rng(OPEX_OFF)})")
    ws.cell(r, TLEASE,  f"=SUM({rng(LEASE_OFF)})")
    ws.cell(r, PRECOMP, f"={cl(TREV)}{r}-{cl(TOPEX)}{r}-{cl(TLEASE)}{r}")
    ws.cell(r, SALYR,   f"=IF({cl(NACTIVE)}{r}=0,0,FLOOR((A{r}-openMo_1)/12,1))")
    ws.cell(r, CORPSAL,
        f"=IF({cl(NACTIVE)}{r}=0,0,((salaryBase+salaryStep*({cl(NACTIVE)}{r}-1))/12)*(1+opexEscalatorPct)^MAX(0,{cl(SALYR)}{r}))")
    ws.cell(r, POSTSAL, f"={cl(PRECOMP)}{r}-{cl(CORPSAL)}{r}")
    ws.cell(r, TDEBT,   f"=SUM({rng(DEBT_OFF)})")
    ws.cell(r, POSTDEBT,f"={cl(POSTSAL)}{r}-{cl(TDEBT)}{r}")
    ws.cell(r, DISTNOI, f"=IF({cl(NACTIVE)}{r}=0,0,({cl(NDIST)}{r}/{cl(NACTIVE)}{r})*{cl(POSTDEBT)}{r})")
    ws.cell(r, PSHARE,  f"=MAX(0,{cl(DISTNOI)}{r})*profitSharePct")
    ws.cell(r, DIST,    f"=MAX(0,{cl(DISTNOI)}{r}-{cl(PSHARE)}{r})")
    # Capital call: sum of locations whose call month is m
    cc_terms = "+".join(
        f"IF(AND(locActive_{j}=1,A{r}=callMo_{j}),investorEquityPerLoc,0)"
        for j in range(1, NLOC+1)
    )
    ws.cell(r, CALL,    f"=-({cc_terms})")
    # Exit: only on the hold-end month
    ws.cell(r, EXITP,   f"=IF(A{r}=holdMonths,exitMultiple*SUMPRODUCT((Month_T12)*({cl(PRECOMP)}2:{cl(PRECOMP)}{MAX_MONTHS+1}))*profitSharePct,0)")
    # Simpler approach for T12: store exit on last row, compute via OFFSET sum
    # We'll handle exit on last month only via dedicated formula:
    if m == MAX_MONTHS:
        # do nothing here — the special exit logic gets a single formula in the next pass
        pass
    ws.cell(r, EXIT,    f"=0")  # placeholder, fixed below
    ws.cell(r, NETCF,   f"={cl(DIST)}{r}+{cl(CALL)}{r}+{cl(EXIT)}{r}")

# Fix exit formulas — only on row where m == holdMonths
# Net equity = MAX(0, gross exit - debt payoff balloon). Promote applied to net.
for m in range(1, MAX_MONTHS + 1):
    r = m + 1
    gross = (f'exitMultiple*SUMPRODUCT(({cl(PRECOMP)}2:{cl(PRECOMP)}{MAX_MONTHS+1})'
             f'*(A2:A{MAX_MONTHS+1}>holdMonths-12)*(A2:A{MAX_MONTHS+1}<=holdMonths))')
    netEquity = f'MAX(0,{gross}-SUM(debtPayoff))'
    ws.cell(r, EXIT,
        f'=IF(A{r}=holdMonths,{netEquity}*(1-profitSharePct),0)')
    ws.cell(r, EXITP,
        f'=IF(A{r}=holdMonths,{netEquity}*profitSharePct,0)')
    ws.cell(r, NETCF,
        f"={cl(DIST)}{r}+{cl(CALL)}{r}+{cl(EXIT)}{r}")

# Number formats on key columns
for m in range(1, MAX_MONTHS + 1):
    r = m + 1
    for cidx in [TREV, TOPEX, TLEASE, PRECOMP, CORPSAL, POSTSAL, TDEBT, POSTDEBT,
                 DISTNOI, PSHARE, DIST, CALL, EXIT, EXITP, NETCF]:
        ws.cell(r, cidx).number_format = USD0

# Named ranges for KPI sheet
named(wb, "M_Month",      f"'Monthly Cash Flow'!$A$2:$A${MAX_MONTHS+1}")
named(wb, "M_NetCF",      f"'Monthly Cash Flow'!${cl(NETCF)}$2:${cl(NETCF)}${MAX_MONTHS+1}")
named(wb, "M_Distrib",    f"'Monthly Cash Flow'!${cl(DIST)}$2:${cl(DIST)}${MAX_MONTHS+1}")
named(wb, "M_Call",       f"'Monthly Cash Flow'!${cl(CALL)}$2:${cl(CALL)}${MAX_MONTHS+1}")
named(wb, "M_Exit",       f"'Monthly Cash Flow'!${cl(EXIT)}$2:${cl(EXIT)}${MAX_MONTHS+1}")
named(wb, "M_NDist",      f"'Monthly Cash Flow'!${cl(NDIST)}$2:${cl(NDIST)}${MAX_MONTHS+1}")
named(wb, "M_NActive",    f"'Monthly Cash Flow'!${cl(NACTIVE)}$2:${cl(NACTIVE)}${MAX_MONTHS+1}")

# ═════ Sheet 12: KPIs ═════
ws = wb.create_sheet("KPIs")
setw(ws, [40, 18])
ws['A1'] = "INVESTOR KPIs"; ws['A1'].font = Font(bold=True, size=14, color="5C4033")

# We rely on Excel XIRR but need date stamps. Build a synthetic date column with EOMONTH.
# Easier: use the IRR function on monthly cash flows, then annualize: (1+monthlyIRR)^12 - 1

rows = [
    ("Total Equity Contributed", "=ABS(SUMIF(M_Month,\"<=\"&holdMonths,M_Call))", USD0),
    ("Total Distributions",       "=SUMIF(M_Month,\"<=\"&holdMonths,M_Distrib)", USD0),
    ("Exit Proceeds (net of promote)", "=SUMIF(M_Month,\"<=\"&holdMonths,M_Exit)", USD0),
    ("Total Returns",             "=B3+B4", USD0),
    ("ROI",                        "=IF(B2>0,(B5-B2)/B2,NA())", PCT1),
    ("MOIC",                       "=IF(B2>0,B5/B2,NA())", MULT),
    # Monthly IRR via IRR on the truncated cash flows; annualize to (1+r)^12-1
    ("Monthly IRR (raw)",          f"=IFERROR(IRR(INDEX(M_NetCF,1):INDEX(M_NetCF,holdMonths)),NA())", PCT2),
    ("Annualized IRR",             "=IF(ISNUMBER(B8),(1+B8)^12-1,NA())", PCT2),
    ("Stabilized CoC (T12 dist / equity)",
        "=IF(AND(B2>0,INDEX(M_NDist,holdMonths)=INDEX(M_NActive,holdMonths),INDEX(M_NActive,holdMonths)>0),"
        "SUMPRODUCT((M_Month>holdMonths-12)*(M_Month<=holdMonths)*M_Distrib)/B2,NA())", PCT2),
    ("Avg CoC (annualized)",
        "=IF(AND(B2>0,holdMonths>0,B3>0),B3/B2/(holdMonths/12),NA())", PCT2),
]
for i, (label, formula, fmt) in enumerate(rows, start=2):
    ws.cell(i, 1, label)
    ws.cell(i, 2, formula); ws.cell(i, 2).number_format = fmt
    ws.cell(i, 2).fill = ACCENT_FILL if "MOIC" in label or "IRR" in label or "CoC" in label else DERIVED_FILL

ws.cell(14, 1, "—— Capital Stack Roll-Up ——").font = SUB
ws.cell(15, 1, "Investor Equity per Location"); ws.cell(15, 2, "=investorEquityPerLoc"); ws.cell(15, 2).number_format = USD0
ws.cell(16, 1, "LP per Location"); ws.cell(16, 2, "=lpPerLoc"); ws.cell(16, 2).number_format = USD0
ws.cell(17, 1, "Total Investor Equity (× numLocations)"); ws.cell(17, 2, "=investorEquityPerLoc*numLocations"); ws.cell(17, 2).number_format = USD0
ws.cell(18, 1, "Total LP Investment"); ws.cell(18, 2, "=lpPerLoc*numLocations"); ws.cell(18, 2).number_format = USD0
ws.cell(19, 1, "Total Debt"); ws.cell(19, 2, "=debtUsed*numLocations"); ws.cell(19, 2).number_format = USD0
ws.cell(20, 1, "Total CapEx"); ws.cell(20, 2, "=totalCapex*numLocations"); ws.cell(20, 2).number_format = USD0
ws.cell(21, 1, "Total TI (DPEG)"); ws.cell(21, 2, "=tiTotal*numLocations"); ws.cell(21, 2).number_format = USD0

ws.cell(23, 1, "Notes").font = SUB
ws.cell(24, 1, "Annualized IRR: Excel IRR returns a monthly rate; this cell raises (1+r)^12-1.").font = NOTE_FONT
ws.cell(25, 1, "Stabilized CoC: shows #N/A unless every active location is distributing in the exit month.").font = NOTE_FONT

# Set sheet ordering
order = ["Notes", "Inputs", "Capital Stack", "Vendor Revenue",
         "Gas", "Electric", "Water", "Non-Utility OpEx", "OpEx Summary",
         "Debt Schedule", "Monthly Cash Flow", "KPIs"]
wb._sheets = [wb[name] for name in order]

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
